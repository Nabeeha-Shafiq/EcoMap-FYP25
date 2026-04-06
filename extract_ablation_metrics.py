#!/usr/bin/env python3
"""
Extract Ablation Study Metrics to Excel

Extracts validation metrics from all ablation study folders (GEO + ZENODO)
and populates an Excel file with standardized formatting.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import json
from pathlib import Path

# ============================================================================
# CONFIGURATION
# ============================================================================

WORKSPACE_ROOT = "/media/hp/ADATA HV300/FYP/Student Model"
ZENODO_BASE = f"{WORKSPACE_ROOT}/ZENODO Ablation Study"
GEO_BASE = f"{WORKSPACE_ROOT}/GEO Ablation Study"
EXCEL_PATH = f"{WORKSPACE_ROOT}/Student Model Ablation Studies.xlsx"

# ZENODO Student Ablations
ZENODO_ABLATIONS = [
    {
        'folder': 'STUDENT_UNIFIED_ZENODO_0.6_PCA_ver5',
        'description': '0.6 PCA retention on UNI embeddings (Zenodo)'
    }
]

# GEO Student Ablations
GEO_ABLATIONS = [
    {
        'folder': 'STUDENT_UNIFIED_GEO_0.6_PCA_ver6',
        'description': '0.6 PCA on UNI embeddings (Baseline)'
    },
    {
        'folder': 'CONCH_STUDENT_UNIFIED_GEO_0.6_PCA',
        'description': 'CONCH encoder with 0.6 PCA'
    },
    {
        'folder': 'H-OPTIMUS_STUDENT_UNIFIED_GEO_0.6_PCA',
        'description': 'H-OPTIMUS encoder with 0.6 PCA'
    },
    {
        'folder': 'UNI_CLASS_BALANCE_STUDENT_UNIFIED_GEO_no_PCA',
        'description': 'UNI with class weights + feature matching (NO PCA)'
    },
    {
        'folder': 'STUDENT_DISTILLATION_0.6_PCA',
        'description': 'Standard distillation with 0.6 PCA'
    }
]

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def extract_metrics_from_json(json_path):
    """
    Extract metrics from student_training_results.json
    
    Returns:
        dict: {accuracy, precision, recall, f1, train_loss, val_loss}
    """
    try:
        with open(json_path, 'r') as f:
            data = json.load(f)
        
        metrics = data.get('overall_metrics', {})
        
        # Try to get loss from per_fold data
        per_fold = data.get('per_fold', [])
        train_loss_final = 0.0
        val_loss_final = 0.0
        
        if per_fold:
            # Use the last epoch's loss from the last fold as an estimate
            train_loss_final = per_fold[-1].get('train_loss_final', 0.0) if 'train_loss_final' in per_fold[-1] else 0.0
            val_loss_final = per_fold[-1].get('val_loss_final', 0.0) if 'val_loss_final' in per_fold[-1] else 0.0
        
        return {
            'accuracy': metrics.get('accuracy', 0.0),
            'precision': metrics.get('precision', 0.0),
            'recall': metrics.get('recall', 0.0),
            'f1': metrics.get('f1', 0.0),
            'train_loss': train_loss_final,
            'val_loss': val_loss_final
        }
    except Exception as e:
        print(f"❌ Error reading {json_path}: {e}")
        return None


def extract_metrics_from_teacher_json(json_path):
    """
    Extract metrics from teacher training_summary.json
    
    Returns:
        dict: {accuracy, precision, recall, f1, train_loss, val_loss}
    """
    try:
        with open(json_path, 'r') as f:
            data = json.load(f)
        
        fold_stats = data.get('fold_stats', {})
        
        return {
            'accuracy': fold_stats.get('accuracy_mean', 0.0),
            'precision': 0.0,  # Not available in teacher summary
            'recall': fold_stats.get('accuracy_mean', 0.0),  # Use accuracy as proxy
            'f1': fold_stats.get('f1_mean', 0.0),
            'train_loss': 0.0,  # Not avg'd in summary
            'val_loss': 0.0     # Not avg'd in summary
        }
    except Exception as e:
        print(f"❌ Error reading {json_path}: {e}")
        return None


def process_ablation(base_path, folder_name, description, is_teacher=False):
    """
    Process a single ablation folder and return metrics
    
    Args:
        base_path: Base directory path (GEO or ZENODO)
        folder_name: Name of the ablation folder
        description: Description of the ablation
        is_teacher: Whether this is a teacher or student model
    
    Returns:
        dict: metrics and metadata, or None if failed
    """
    folder_path = os.path.join(base_path, folder_name)
    
    if not os.path.exists(folder_path):
        print(f"⚠️  Folder not found: {folder_path}")
        return None
    
    # Determine JSON file name based on model type
    if is_teacher:
        json_file = 'training_summary.json'
        json_path = os.path.join(folder_path, 'training', 'metrics', json_file)
        if not os.path.exists(json_path):
            print(f"⚠️  {json_file} not found in {folder_path}")
            return None
        metrics = extract_metrics_from_teacher_json(json_path)
    else:
        json_file = 'student_training_results.json'
        json_path = os.path.join(folder_path, 'training', 'metrics', json_file)
        if not os.path.exists(json_path):
            print(f"⚠️  {json_file} not found in {folder_path}")
            return None
        metrics = extract_metrics_from_json(json_path)
    
    if metrics is None:
        return None
    
    return {
        'folder': folder_name,
        'description': description,
        **metrics
    }


def create_excel_workbook():
    """Create a new Excel workbook with proper formatting"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ablation Studies"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_font = Font(bold=True, size=11)
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    
    # Add main header
    ws['A1'] = "Dataset"
    ws['B1'] = "Folder Name"
    ws['C1'] = "Description"
    ws['D1'] = "Accuracy"
    ws['E1'] = "Precision"
    ws['F1'] = "Recall"
    ws['G1'] = "F1-Score"
    ws['H1'] = "Train Loss"
    ws['I1'] = "Val Loss"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].font = header_font
        ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
    
    return wb, ws, header_fill, section_fill, header_font, section_font


def populate_zenodo_data(ws, start_row=2):
    """Populate ZENODO ablation data"""
    current_row = start_row
    
    # Section header
    ws[f'A{current_row}'] = "ZENODO ABLATION STUDIES"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{current_row}'].fill = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
    ws.merge_cells(f'A{current_row}:I{current_row}')
    current_row += 1
    
    # Process each ablation
    for ablation in ZENODO_ABLATIONS:
        result = process_ablation(
            ZENODO_BASE,
            ablation['folder'],
            ablation['description'],
            is_teacher=False
        )
        
        if result:
            ws[f'A{current_row}'] = "ZENODO"
            ws[f'B{current_row}'] = result['folder']
            ws[f'C{current_row}'] = result['description']
            ws[f'D{current_row}'] = round(result['accuracy'], 6)
            ws[f'E{current_row}'] = round(result['precision'], 6)
            ws[f'F{current_row}'] = round(result['recall'], 6)
            ws[f'G{current_row}'] = round(result['f1'], 6)
            ws[f'H{current_row}'] = round(result['train_loss'], 6)
            ws[f'I{current_row}'] = round(result['val_loss'], 6)
            
            # Center alignment for numeric columns
            for col in ['D', 'E', 'F', 'G', 'H', 'I']:
                ws[f'{col}{current_row}'].alignment = Alignment(horizontal='center')
            
            print(f"✓ Zenodo: {ablation['folder']} - Accuracy: {result['accuracy']:.6f}")
            current_row += 1
        else:
            print(f"✗ Failed to process: {ablation['folder']}")
    
    return current_row


def populate_geo_data(ws, start_row):
    """Populate GEO ablation data"""
    # Add blank row
    start_row += 1
    
    # Section header
    ws[f'A{start_row}'] = "GEO ABLATION STUDIES"
    ws[f'A{start_row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{start_row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{start_row}:I{start_row}')
    current_row = start_row + 1
    
    # Process each ablation
    for ablation in GEO_ABLATIONS:
        result = process_ablation(
            GEO_BASE,
            ablation['folder'],
            ablation['description'],
            is_teacher=False
        )
        
        if result:
            ws[f'A{current_row}'] = "GEO"
            ws[f'B{current_row}'] = result['folder']
            ws[f'C{current_row}'] = result['description']
            ws[f'D{current_row}'] = round(result['accuracy'], 6)
            ws[f'E{current_row}'] = round(result['precision'], 6)
            ws[f'F{current_row}'] = round(result['recall'], 6)
            ws[f'G{current_row}'] = round(result['f1'], 6)
            ws[f'H{current_row}'] = round(result['train_loss'], 6)
            ws[f'I{current_row}'] = round(result['val_loss'], 6)
            
            # Center alignment for numeric columns
            for col in ['D', 'E', 'F', 'G', 'H', 'I']:
                ws[f'{col}{current_row}'].alignment = Alignment(horizontal='center')
            
            print(f"✓ GEO: {ablation['folder']} - Accuracy: {result['accuracy']:.6f}")
            current_row += 1
        else:
            print(f"✗ Failed to process: {ablation['folder']}")
    
    return current_row


def verify_excel(excel_path):
    """Verify that data was written to Excel"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Check a few key cells
        sample_checks = [
            ('A1', 'Dataset'),
            ('A2', 'ZENODO ABLATION STUDIES'),
            ('D3', 'Accuracy value'),
        ]
        
        print("\n✓ VERIFICATION:")
        print(f"  Excel file exists: {os.path.exists(excel_path)}")
        print(f"  File size: {os.path.getsize(excel_path)} bytes")
        print(f"  Active sheet: {ws.title}")
        
        # Sample data verification
        if ws['A1'].value == 'Dataset':
            print("  ✓ Headers written correctly")
        
        # Count rows with data
        row_count = 0
        for row in ws.iter_rows(min_row=2, max_row=100, max_col=1, values_only=True):
            if row[0]:
                row_count += 1
        
        print(f"  ✓ Data rows populated: {row_count}")
        
        return True
    except Exception as e:
        print(f"❌ Verification failed: {e}")
        return False


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    print("=" * 80)
    print("EXTRACTING ABLATION STUDY METRICS TO EXCEL")
    print("=" * 80)
    
    # Create workbook
    print("\n[1] Creating Excel workbook...")
    wb, ws, header_fill, section_fill, header_font, section_font = create_excel_workbook()
    
    # Populate ZENODO data
    print("[2] Processing ZENODO ablations...")
    zenodo_end_row = populate_zenodo_data(ws, start_row=2)
    
    # Populate GEO data
    print("[3] Processing GEO ablations...")
    geo_end_row = populate_geo_data(ws, start_row=zenodo_end_row)
    
    # Save workbook
    print(f"\n[4] Saving to {EXCEL_PATH}...")
    try:
        wb.save(EXCEL_PATH)
        print("✓ Excel file saved successfully")
    except Exception as e:
        print(f"❌ Failed to save Excel: {e}")
        return False
    
    # Verify
    print("\n[5] Verifying Excel file...")
    if verify_excel(EXCEL_PATH):
        print("\n" + "=" * 80)
        print("✅ EXTRACTION COMPLETE")
        print(f"   Excel file: {EXCEL_PATH}")
        print("=" * 80)
        return True
    else:
        print("\n❌ Verification failed")
        return False


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
